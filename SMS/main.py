import dgl
import openpyxl
from graph_extraction_unk import get_graph_from_sms

def read_excel():
    workbook = openpyxl.load_workbook('test_dataset.xlsx')
    sheet = workbook.active
    print(sheet.title, sheet.max_row, sheet.max_column)

    don_sms_col = [cell.value for cell in sheet['E'][1:]]
    acc_sms_col = [cell.value for cell in sheet['D'][1:]]

    print(len(don_sms_col),len(acc_sms_col))

    return don_sms_col, acc_sms_col

def main():
    don_sms_col, acc_sms_col = read_excel()

    don_graphs = []
    for i in range(len(don_sms_col)):
        try:
            don_tmp = get_graph_from_sms(don_sms_col[i])
            if don_tmp.num_nodes() != 0 or don_tmp.num.edges() != 0 :
                don_graphs.append(don_tmp)
        except Exception as e:
            print(i)
            pass
    print(len(don_graphs))
    
    acc_graphs = []
    for i in range(len(acc_sms_col)):
        try:
            acc_tmp = get_graph_from_sms(acc_sms_col[i])
            if acc_tmp.num_nodes() != 0 and acc_tmp.num_edges() != 0 :
                acc_graphs.append(acc_tmp)
        except:
            print(i)
            pass
    print(len(acc_graphs))

    
    dgl.save_graphs('don.dgl',don_graphs)
    dgl.save_graphs('acc.dgl',acc_graphs)
    dg, _ = dgl.load_graphs('don.dgl')
    ag, _ = dgl.load_graphs('acc.dgl')
    
    print(len(dg),len(ag))

if __name__ == '__main__':
    main()

